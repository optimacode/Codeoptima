import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("gmail_sektoru.xlsx")
ws = wb.active
while True:
  # Kullanıcıya seçenekler sunma
  print("1 - Yeni satış ekle")
  print("2 - Verileri güncelle")
  print("3 - Verileri ara")
  print("4 - Çıkış yap")
  secim = input("Seçiminiz: ")

  # Yeni satış ekleme
  if secim == "1":
    # Kullanıcıdan veri girişi alma
    ad = input("Adınızı girin: ")
    soyad = input("Soyadınızı girin: ")
    r10_kullanici_adi = input("R10 kullanıcı adınızı girin: ")
    miktar = float(input("Miktarı girin: "))
    maliyet = float(input("Maliyeti girin: "))
    alinan_tutar = float(input("Alınan tutarı girin: "))
    proxy_adet = float(input("Kaç Adet Proxy Kullanıldı: "))
    proxy_maliyeti = float(input("Proxy Maliyeti:"))
    dolar_kuru = float(input("Dolar Kuru:"))
    sms = float(input("Kaç adet sms kullanıldı:"))
    sms_maliyet = float(input("Bir adet sms maliyeti:"))
    sonuc = proxy_maliyeti * proxy_adet
    cevir = sonuc * dolar_kuru
    sms_dolar_bazlı = sms_maliyet * sms
    sms_cevir = sms_dolar_bazlı * dolar_kuru
    tarih = datetime.today().strftime('%Y-%m-%d')
    hesap_durumu = "Aktif"
    kazanc = alinan_tutar - maliyet

    # Add new row to the worksheet
    ws.append([
      ad, soyad, r10_kullanici_adi, miktar, maliyet, alinan_tutar, kazanc,
      tarih, hesap_durumu, proxy_maliyeti, proxy_adet, cevir, sms, sms_maliyet,
      sms_cevir
    ])

    # Save the workbook
    wb.save("gmail_sektoru.xlsx")

  # Verileri güncelleme
  elif secim == "2":
    # Kullanıcıdan R10 kullanıcı adını alma
    r10_kullanici_adi = input(
      "Güncellemek istediğiniz hesabın R10 kullanıcı adını girin: ")

    for row in ws.iter_rows(min_row=2, values_only=True):
      if row[2] == r10_kullanici_adi:
        # Yeni verileri kullanıcıdan alma
        miktar = float(input("Miktarı girin: "))
        maliyet = float(input("Maliyeti girin: "))
        alinan_tutar = float(input("Alınan tutarı girin: "))
        hesap_durumu = input("Hesap durumunu girin: ")
        proxy_maliyeti = float(input("Proxy Maliyeti:"))
        proxy_adet = float(input("Proxy Adet:"))
        dolar_kuru = float(input("Dolar Kuru:"))
        sms = float(input("Kaç adet sms kullanıldı:"))
        sms_maliyet = float(input("Bir adet sms maliyeti:"))
        sonuc = proxy_maliyeti * proxy_adet
        cevir = sonuc * dolar_kuru
        sms_dolar_bazlı = sms_maliyet * sms
        sms_cevir = sms_dolar_bazlı * dolar_kuru
        kazanc = alinan_tutar - maliyet
        tarih = datetime.today().strftime('%Y-%m-%d')

        # Update the row
        row = [
          row[0], row[1], row[2], miktar, maliyet, alinan_tutar, kazanc, tarih,
          hesap_durumu, proxy_maliyeti, proxy_adet, cevir, sms, sms_maliyet,
          sms_cevir
        ]
        ws.append(row)

    # Save the workbook
    wb.save("gmail_sektoru.xlsx")

  # Veri arama
  elif secim == "3":
    # Kullanıcıdan aranacak R10 kullanıcı adını alma
    aranan_kullanici_adi = input("Aranacak R10 kullanıcı adını girin: ")

    for row in ws.iter_rows(min_row=2, values_only=True):
      if row[2] == aranan_kullanici_adi:
        print(f"{row[0]} {row[1]} hesabı bulundu.")
        print(
          f"Miktar: {row[3]}, Maliyet: {row[4]}, Alınan Tutar: {row[5]}, Kazanç: {row[6]}, Tarih: {row[7]}, Durum: {row[8]}, Proxy Maliyeti: {row[9]}, Proxy Adet: {row[10]}, Proxy Net Maliyet: {row[11]}, Sms Adet: {row[12]},Sms Maliyet: {row[13]}, Sms Dolar Bazlı Maliyet: {row[14]}"
        )
        break
    else:
      print(f"{aranan_kullanici_adi} R10 kullanıcısı bulunamadı.")
  elif secim == "4":
    print("Programdan çıkılıyor...")
    break
